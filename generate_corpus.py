# 3열이 비어 있으면 먼저 전체를 포스태깅하고, txt로 변환
# (표준어\t제주어\t표준어포스태깅\t제주어포스태깅\n)
# python generate_corpus [Copora Directory] [Output Directory]

import os
import sys
import openpyxl

from konlpy.tag import Okt


def main(corpora, output):
    filelist = os.listdir(corpora)
    tagger = Okt()

    for file in filelist:
        book = openpyxl.load_workbook(os.path.join(corpora, file))
        sheet = book.get_sheet_by_name("Sheet1")

        tagged = (bool(sheet.cell(row=1, column=3).value) and bool(sheet.cell(row=1, column=4).value))

        if not tagged:
            for sample in sheet.rows:
                index = sample[0].row

                stan = sample[0].value
                pos_stan = '/'.join(tagger.morphs(stan))
                jeju = sample[1].value
                pos_jeju = '/'.join(tagger.morphs(jeju))

                sheet.cell(row=index, column=3).value = pos_stan
                sheet.cell(row=index, column=4).value = pos_jeju

            book.save(os.path.join(corpora, file))

        filename = file[:file.find('.')]
        output_dir = os.path.join(output, filename + '.txt')    # Exception: Output Dir not Exists
        output_file = open(output_dir, 'w')

        for sample in sheet.rows:
            line = '\t'.join([s.value for s in sample[:5]]) + '\n'  # Exception: s.value can be no string
            output_file.write(line)

        output_file.close()
        book.close()


if __name__ == '__main__':
    args = sys.argv[1:]
    main(*args)
